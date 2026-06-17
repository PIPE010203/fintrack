from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib import messages
from django.db.models import Sum
from .models import Gasto, Ingreso
from .forms import GastoForm, IngresoForm
from presupuesto.models import Presupuesto
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

MESES = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
    5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
}

@login_required
def lista_gastos(request):
    gastos = Gasto.objects.filter(usuario=request.user)
    hoy = date.today()
    presupuestos = Presupuesto.objects.filter(
        usuario=request.user, mes=hoy.month, anio=hoy.year
    )
    alertas = []
    for p in presupuestos:
        gastado = Gasto.objects.filter(
            usuario=request.user,
            categoria=p.categoria,
            fecha__month=hoy.month,
            fecha__year=hoy.year
        ).aggregate(total=Sum('monto'))['total'] or 0
        if float(gastado) > float(p.limite):
            exceso = float(gastado) - float(p.limite)
            alertas.append({
                'categoria': p.get_categoria_display(),
                'limite': p.limite,
                'gastado': gastado,
                'exceso': exceso,
            })
    return render(request, 'transacciones/gastos.html', {
        'gastos': gastos,
        'alertas': alertas,
    })

@login_required
def nuevo_gasto(request):
    form = GastoForm(request.POST or None)
    if form.is_valid():
        g = form.save(commit=False)
        g.usuario = request.user
        g.save()
        messages.success(request, 'Gasto registrado.')
        return redirect('gastos')
    return render(request, 'transacciones/form_gasto.html', {'form': form, 'titulo': 'Nuevo Gasto'})

@login_required
def editar_gasto(request, pk):
    gasto = get_object_or_404(Gasto, pk=pk, usuario=request.user)
    form = GastoForm(request.POST or None, instance=gasto)
    if form.is_valid():
        form.save()
        messages.success(request, 'Gasto actualizado.')
        return redirect('gastos')
    return render(request, 'transacciones/form_gasto.html', {'form': form, 'titulo': 'Editar Gasto'})

@login_required
def eliminar_gasto(request, pk):
    gasto = get_object_or_404(Gasto, pk=pk, usuario=request.user)
    if request.method == 'POST':
        gasto.delete()
        messages.success(request, 'Gasto eliminado.')
        return redirect('gastos')
    return render(request, 'transacciones/confirmar_eliminar.html', {'objeto': gasto, 'tipo': 'gasto'})

@login_required
def lista_ingresos(request):
    ingresos = Ingreso.objects.filter(usuario=request.user)
    return render(request, 'transacciones/ingresos.html', {'ingresos': ingresos})

@login_required
def nuevo_ingreso(request):
    form = IngresoForm(request.POST or None)
    if form.is_valid():
        i = form.save(commit=False)
        i.usuario = request.user
        i.save()
        messages.success(request, 'Ingreso registrado.')
        return redirect('ingresos')
    return render(request, 'transacciones/form_ingreso.html', {'form': form, 'titulo': 'Nuevo Ingreso'})

@login_required
def editar_ingreso(request, pk):
    ingreso = get_object_or_404(Ingreso, pk=pk, usuario=request.user)
    form = IngresoForm(request.POST or None, instance=ingreso)
    if form.is_valid():
        form.save()
        messages.success(request, 'Ingreso actualizado.')
        return redirect('ingresos')
    return render(request, 'transacciones/form_ingreso.html', {'form': form, 'titulo': 'Editar Ingreso'})

@login_required
def eliminar_ingreso(request, pk):
    ingreso = get_object_or_404(Ingreso, pk=pk, usuario=request.user)
    if request.method == 'POST':
        ingreso.delete()
        messages.success(request, 'Ingreso eliminado.')
        return redirect('ingresos')
    return render(request, 'transacciones/confirmar_eliminar.html', {'objeto': ingreso, 'tipo': 'ingreso'})

@login_required
def selector_reporte(request):
    hoy = date.today()
    anios = list(range(2020, hoy.year + 1))
    return render(request, 'transacciones/selector_reporte.html', {
        'meses': MESES,
        'anios': anios,
        'mes_actual': hoy.month,
        'anio_actual': hoy.year,
    })

@login_required
def exportar_pdf(request):
    hoy = date.today()
    mes = request.GET.get('mes', '')
    anio = request.GET.get('anio', str(hoy.year))
    todo = request.GET.get('todo', '')

    gastos = Gasto.objects.filter(usuario=request.user).order_by('fecha')
    ingresos = Ingreso.objects.filter(usuario=request.user).order_by('fecha')

    if todo:
        titulo_periodo = "Historial completo"
    else:
        mes = int(mes) if mes else hoy.month
        anio = int(anio)
        gastos = gastos.filter(fecha__month=mes, fecha__year=anio)
        ingresos = ingresos.filter(fecha__month=mes, fecha__year=anio)
        titulo_periodo = f"{MESES[mes]} {anio}"

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="fintrack_{titulo_periodo}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4,
                           rightMargin=2*cm, leftMargin=2*cm,
                           topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    elements = []

    # Título
    titulo_style = ParagraphStyle('titulo', fontSize=20, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#1a1a2e'), spaceAfter=5)
    sub_style = ParagraphStyle('sub', fontSize=11, textColor=colors.HexColor('#666666'), spaceAfter=20)

    elements.append(Paragraph("FinTrack — Reporte Financiero", titulo_style))
    elements.append(Spacer(1, 0.3*cm))
    elements.append(Paragraph(f"Usuario: {request.user.username}   |   Período: {titulo_periodo}   |   Generado: {hoy.strftime('%d/%m/%Y')}", sub_style))
    elements.append(Spacer(1, 0.5*cm))

    # Colores
    color_header_gasto = colors.HexColor('#c0392b')
    color_header_ingreso = colors.HexColor('#27ae60')
    color_fila_par = colors.HexColor('#f9f9f9')
    color_blanco = colors.white
    color_total = colors.HexColor('#1a1a2e')

    def hacer_tabla(datos, encabezado, color_header):
        tabla_data = [encabezado] + datos
        tabla = Table(tabla_data, colWidths=[3*cm, 6*cm, 4*cm, 4*cm])
        estilo = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), color_header),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ALIGN', (1,1), (1,-1), 'LEFT'),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [color_blanco, color_fila_par]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ])
        tabla.setStyle(estilo)
        return tabla

    # GASTOS
    section_style = ParagraphStyle('section', fontSize=13, fontName='Helvetica-Bold',
                                   textColor=colors.HexColor('#c0392b'), spaceAfter=8, spaceBefore=15)
    elements.append(Paragraph("GASTOS", section_style))

    total_gastos = 0
    filas_gastos = []
    for g in gastos:
        filas_gastos.append([
            g.fecha.strftime('%d/%m/%Y'),
            g.descripcion.capitalize(),
            g.get_categoria_display(),
            f"${float(g.monto):,.0f} COP"
        ])
        total_gastos += float(g.monto)

    if filas_gastos:
        elements.append(hacer_tabla(filas_gastos, ['Fecha', 'Descripción', 'Categoría', 'Monto'], color_header_gasto))
    else:
        elements.append(Paragraph("Sin gastos en este período.", styles['Normal']))

    elements.append(Spacer(1, 0.3*cm))
    total_style = ParagraphStyle('total', fontSize=11, fontName='Helvetica-Bold',
                                 textColor=color_total, spaceAfter=5)
    elements.append(Paragraph(f"Total gastos: ${total_gastos:,.0f} COP", total_style))

    # INGRESOS
    section_style2 = ParagraphStyle('section2', fontSize=13, fontName='Helvetica-Bold',
                                    textColor=colors.HexColor('#27ae60'), spaceAfter=8, spaceBefore=15)
    elements.append(Paragraph("INGRESOS", section_style2))

    total_ingresos = 0
    filas_ingresos = []
    for i in ingresos:
        filas_ingresos.append([
            i.fecha.strftime('%d/%m/%Y'),
            i.descripcion.capitalize(),
            i.get_categoria_display(),
            f"${float(i.monto):,.0f} COP"
        ])
        total_ingresos += float(i.monto)

    if filas_ingresos:
        elements.append(hacer_tabla(filas_ingresos, ['Fecha', 'Descripción', 'Categoría', 'Monto'], color_header_ingreso))
    else:
        elements.append(Paragraph("Sin ingresos en este período.", styles['Normal']))

    elements.append(Spacer(1, 0.3*cm))
    elements.append(Paragraph(f"Total ingresos: ${total_ingresos:,.0f} COP", total_style))

    balance = total_ingresos - total_gastos
    color_balance = colors.HexColor('#27ae60') if balance >= 0 else colors.HexColor('#c0392b')
    balance_style = ParagraphStyle('balance', fontSize=13, fontName='Helvetica-Bold',
                                   textColor=color_balance, spaceBefore=10)
    elements.append(Paragraph(f"Balance: ${balance:,.0f} COP", balance_style))

    doc.build(elements)
    return response

@login_required
def exportar_excel(request):
    hoy = date.today()
    mes = request.GET.get('mes', '')
    anio = request.GET.get('anio', str(hoy.year))
    todo = request.GET.get('todo', '')

    gastos = Gasto.objects.filter(usuario=request.user).order_by('fecha')
    ingresos = Ingreso.objects.filter(usuario=request.user).order_by('fecha')

    if todo:
        titulo_periodo = "Historial completo"
    else:
        mes = int(mes) if mes else hoy.month
        anio = int(anio)
        gastos = gastos.filter(fecha__month=mes, fecha__year=anio)
        ingresos = ingresos.filter(fecha__month=mes, fecha__year=anio)
        titulo_periodo = f"{MESES[mes]} {anio}"

    wb = openpyxl.Workbook()

    # Estilos
    header_gasto = PatternFill("solid", fgColor="C0392B")
    header_ingreso = PatternFill("solid", fgColor="27AE60")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14, color="1A1A2E")
    fila_par = PatternFill("solid", fgColor="F9F9F9")
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    center = Alignment(horizontal='center', vertical='center')

    def crear_hoja(ws, titulo, filas, encabezados, fill_header, total, tipo):
        ws.append([f"FinTrack — {titulo}"])
        ws['A1'].font = title_font
        ws.append([f"Usuario: {request.user.username}   |   Período: {titulo_periodo}"])
        ws.append([])
        ws.append(encabezados)
        for col, cell in enumerate(ws[4], 1):
            cell.fill = fill_header
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for i, fila in enumerate(filas, 5):
            ws.append(fila)
            for cell in ws[i]:
                cell.border = border
                cell.alignment = center
                if i % 2 == 0:
                    cell.fill = fila_par

        ws.append([])
        total_row = ws.max_row + 1
        ws.append([f"Total {tipo}:", '', '', f"${total:,.0f} COP"])
        ws[f'A{total_row}'].font = Font(bold=True, size=11)
        ws[f'D{total_row}'].font = Font(bold=True, size=11)

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

    # Hoja gastos
    ws1 = wb.active
    ws1.title = "Gastos"
    total_gastos = 0
    filas_g = []
    for g in gastos:
        filas_g.append([g.fecha.strftime('%d/%m/%Y'), g.descripcion.capitalize(),
                        g.get_categoria_display(), f"${float(g.monto):,.0f} COP"])
        total_gastos += float(g.monto)
    crear_hoja(ws1, "Gastos", filas_g, ['Fecha', 'Descripción', 'Categoría', 'Monto'],
               header_gasto, total_gastos, "gastos")

    # Hoja ingresos
    ws2 = wb.create_sheet("Ingresos")
    total_ingresos = 0
    filas_i = []
    for i in ingresos:
        filas_i.append([i.fecha.strftime('%d/%m/%Y'), i.descripcion.capitalize(),
                        i.get_categoria_display(), f"${float(i.monto):,.0f} COP"])
        total_ingresos += float(i.monto)
    crear_hoja(ws2, "Ingresos", filas_i, ['Fecha', 'Descripción', 'Categoría', 'Monto'],
               header_ingreso, total_ingresos, "ingresos")

    # Hoja resumen
    ws3 = wb.create_sheet("Resumen")
    ws3.append(["FinTrack — Resumen Financiero"])
    ws3['A1'].font = title_font
    ws3.append([f"Período: {titulo_periodo}"])
    ws3.append([])
    ws3.append(['Concepto', 'Monto'])
    ws3['A4'].font = Font(bold=True); ws3['B4'].font = Font(bold=True)
    balance = total_ingresos - total_gastos
    ws3.append(['Total Ingresos', f"${total_ingresos:,.0f} COP"])
    ws3.append(['Total Gastos', f"${total_gastos:,.0f} COP"])
    ws3.append(['Balance', f"${balance:,.0f} COP"])
    ws3['A7'].font = Font(bold=True, size=12)
    ws3['B7'].font = Font(bold=True, size=12,
                          color="27AE60" if balance >= 0 else "C0392B")
    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="fintrack_{titulo_periodo}.xlsx"'
    wb.save(response)
    return response